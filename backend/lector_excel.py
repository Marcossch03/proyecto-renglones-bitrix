from pathlib import Path
from openpyxl import load_workbook


HOJA_PRECIOS = "PLANILLA DE PRECIOS"

# Según la planilla modelo:
# A = Renglón
# B = Descripción
# C = Cantidad
# D = Valor unitario
# E = Valor total
# H a K = dotación 1h a 4h
# L a P = dotación mayor a 4h
# Q = Supervisor
COLUMNAS = {
    "renglon": "A",
    "descripcion": "B",
    "cantidad": "C",
    "valor_unitario": "D",
    "valor_total": "E",
    "hora_1": "H",
    "hora_2": "I",
    "hora_3": "J",
    "hora_4": "K",
    "hora_5": "L",
    "hora_6": "M",
    "hora_7": "N",
    "hora_8": "O",
    "hora_9": "P",
    "supervisor": "Q",
    "empresa_actual": "S",
    "dotacion_texto": "T",
    "frecuencia": "U",
    "domicilio": "V",
}


def limpiar_valor(valor):
    """
    Normaliza valores vacíos de Excel.
    """
    if valor is None:
        return None

    if isinstance(valor, str):
        valor = valor.strip()
        if valor == "":
            return None

    return valor


def convertir_numero(valor):
    """
    Convierte valores numéricos de Excel.
    Si viene vacío o con texto no numérico, devuelve 0.
    """
    valor = limpiar_valor(valor)

    if valor is None:
        return 0

    if isinstance(valor, (int, float)):
        return valor

    if isinstance(valor, str):
        valor = valor.replace("$", "").strip()
        valor = valor.replace(".", "").replace(",", ".")

        try:
            return float(valor)
        except ValueError:
            return 0

    return 0


def obtener_valor(hoja, columna, fila):
    return limpiar_valor(hoja[f"{columna}{fila}"].value)


def calcular_dotaciones(hoja, fila):
    """
    Regla definida:
    - De 0 a 4 hs => operario de 4hs
    - Más de 4 hs => operario de 8hs

    El supervisor se lee aparte y NO se suma a la dotación.
    """

    dotacion_4hs = 0
    dotacion_8hs = 0

    columnas_4hs = [
        COLUMNAS["hora_1"],
        COLUMNAS["hora_2"],
        COLUMNAS["hora_3"],
        COLUMNAS["hora_4"],
    ]

    columnas_8hs = [
        COLUMNAS["hora_5"],
        COLUMNAS["hora_6"],
        COLUMNAS["hora_7"],
        COLUMNAS["hora_8"],
        COLUMNAS["hora_9"],
    ]

    for columna in columnas_4hs:
        dotacion_4hs += convertir_numero(hoja[f"{columna}{fila}"].value)

    for columna in columnas_8hs:
        dotacion_8hs += convertir_numero(hoja[f"{columna}{fila}"].value)

    return dotacion_4hs, dotacion_8hs


def fila_es_renglon_valido(hoja, fila):
    """
    Consideramos válida una fila solo si el campo renglón es numérico.
    Esto evita tomar filas de totales, títulos o comentarios.
    """
    renglon = obtener_valor(hoja, COLUMNAS["renglon"], fila)

    if renglon is None:
        return False

    if isinstance(renglon, int):
        return True

    if isinstance(renglon, float) and renglon.is_integer():
        return True

    if isinstance(renglon, str):
        return renglon.strip().isdigit()

    return False


def leer_renglones(ruta_excel):
    ruta_excel = Path(ruta_excel)

    if not ruta_excel.exists():
        raise FileNotFoundError(f"No existe el archivo: {ruta_excel}")

    workbook = load_workbook(ruta_excel, data_only=True)

    if HOJA_PRECIOS not in workbook.sheetnames:
        raise ValueError(f"No se encontró la hoja '{HOJA_PRECIOS}'")

    hoja = workbook[HOJA_PRECIOS]

    renglones = []

    # En la planilla modelo los renglones empiezan en la fila 4.
    fila_inicio = 4

    for fila in range(fila_inicio, hoja.max_row + 1):
        if not fila_es_renglon_valido(hoja, fila):
            continue

        dotacion_4hs, dotacion_8hs = calcular_dotaciones(hoja, fila)

        renglon = {
            "fila_excel": fila,

            # Mapeo futuro a Bitrix:
            # nombre_producto -> NAME
            # descripcion_producto -> DESCRIPTION
            "nombre_producto": str(obtener_valor(hoja, COLUMNAS["renglon"], fila)),
            "descripcion_producto": obtener_valor(hoja, COLUMNAS["descripcion"], fila),

            # Datos económicos
            "cantidad": convertir_numero(obtener_valor(hoja, COLUMNAS["cantidad"], fila)),
            "valor_unitario": convertir_numero(obtener_valor(hoja, COLUMNAS["valor_unitario"], fila)),
            "valor_total": convertir_numero(obtener_valor(hoja, COLUMNAS["valor_total"], fila)),

            # Datos operativos
            "dotacion_4hs": dotacion_4hs,
            "dotacion_8hs": dotacion_8hs,
            "supervisor": convertir_numero(obtener_valor(hoja, COLUMNAS["supervisor"], fila)),

            # Datos informativos
            "empresa_actual": obtener_valor(hoja, COLUMNAS["empresa_actual"], fila),
            "dotacion_texto": obtener_valor(hoja, COLUMNAS["dotacion_texto"], fila),
            "frecuencia": obtener_valor(hoja, COLUMNAS["frecuencia"], fila),
            "domicilio": obtener_valor(hoja, COLUMNAS["domicilio"], fila),
        }

        renglones.append(renglon)

    return renglones


def mostrar_vista_previa(renglones):
    print("\n=== Vista previa de renglones detectados ===")
    print(f"Cantidad de renglones detectados: {len(renglones)}\n")

    for item in renglones:
        print(
            f"Fila {item['fila_excel']} | "
            f"Producto: {item['nombre_producto']} | "
            f"Descripción: {item['descripcion_producto']} | "
            f"Valor unitario: {item['valor_unitario']} | "
            f"Valor total: {item['valor_total']} | "
            f"Dotación 4hs: {item['dotacion_4hs']} | "
            f"Dotación 8hs: {item['dotacion_8hs']} | "
            f"Supervisores: {item['supervisor']}"
        )